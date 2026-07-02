#!/usr/bin/env python3
from __future__ import annotations

from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET
from zipfile import ZipFile


ROOT = Path(__file__).resolve().parents[2]
QA = ROOT / "qa"
WORKBOOK = QA / "TESTING_TEST_CASES.xlsx"
TIMELINE_LOG = QA / "QA_TIMELINE_VALIDATION_LOG.md"
LEGACY_LOG = QA / "QA_SHEET_VALIDATION_LOG.md"

START_DATE = "2026-06-10"
START_TIME = "09:00 IST"
COMPLETION_DATE = "2026-06-16"
FINAL_SUBMISSION = "2026-06-16 18:00 IST"

FILES_CHANGED = [
    "qa/TESTING_TEST_CASES.xlsx",
    "qa/TESTING_TEST_CASES.xlsx: How To Use and Tester Guide sheets",
    "qa/TESTING_CHECKLIST_INTERNAL.md",
    "qa/TESTING_CHECKLIST_CLIENT.md",
    "qa/TESTING_CHECKLIST_CLIENT.docx",
    "qa/TESTER_RUN_BOOK.md",
    "qa/RELEASE_SIGNOFF_SUMMARY.md",
    "qa/QA_TIMELINE_EXECUTION_PLAN.md",
    "qa/QA_DAILY_SUBMISSION_PLAN.md",
    "qa/QA_TIMELINE_VALIDATION_LOG.md",
    "qa/QA_SHEET_VALIDATION_LOG.md",
    "qa/scripts/validate_testing_workbook.py",
    "qa/TASK_SHEET_UPDATE_PATCH.md",
    "qa/EVIDENCE_REGISTER.md",
    "docs/implementation/HRMS_PRODUCTION_TASK_SHEET.md",
    "hrms_backend/docs/implementation/HRMS_PRODUCTION_TASK_SHEET.md",
    "qa/scripts/generate_testing_artifacts.py (removed)",
    "qa/scripts/generate_full_qa_artifacts.py (removed)",
    "qa/scripts/generate_deployment_agile_artifacts.py (removed)",
    "qa/scripts/validate_qa_artifacts.py (removed)",
    "qa/QA_ARTIFACT_UPGRADE_ANALYSIS.md (removed)",
    "qa/QA_ARTIFACT_VALIDATION_LOG.md (removed)",
    "qa/QA_ARTIFACT_DEV_TEST_LOG.md (removed)",
    "qa/DEPLOYMENT_AGILE_IMPLEMENTATION_LOG.md (removed)",
    "qa/DEV_TEST_EXECUTION_LOG.md (removed)",
]

REQUIRED_SHEETS = [
    "Executive Summary",
    "How To Use",
    "Environment Matrix",
    "Agile Sprint Plan",
    "Main Test Cases",
    "Execution Summary",
    "Release Gates",
    "P0 Smoke UAT Gate",
    "Deployment Smoke",
    "CI-CD Validation",
    "Domain DNS SSL CORS",
    "Data Isolation",
    "Secrets Config Checklist",
    "Branch Promotion Matrix",
    "Sprint Regression",
    "Full Regression",
    "Role Permission Matrix",
    "Business Rule Traceability",
    "Test Data",
    "Defect Log",
    "Rollback Checklist",
    "Future Scope",
    "Signoff",
]

EXECUTION_SHEETS = [
    "P0 Smoke UAT Gate",
    "Deployment Smoke",
    "CI-CD Validation",
    "Domain DNS SSL CORS",
    "Data Isolation",
    "Secrets Config Checklist",
    "Branch Promotion Matrix",
    "Sprint Regression",
    "Full Regression",
    "Rollback Checklist",
]

CASE_COLUMNS = [
    "Test Case ID",
    "Sprint",
    "Execution Day",
    "Planned Start Date",
    "Planned End Date",
    "Daily Submission Due",
    "Final Submission Due",
    "Completion Owner",
    "Reviewer / Signoff Owner",
    "Execution Lane",
    "Priority",
    "Story Points",
    "Status",
    "Actual Result",
    "Evidence Required",
    "Evidence Link / Screenshot Ref",
    "Defect ID",
    "Blocked Reason",
    "Retest Due Date",
    "Submission Status",
]

LANES = {
    "P0 Release Gate",
    "Deployment Smoke",
    "Core HR Regression",
    "Finance/Expense Regression",
    "Documents/Media Regression",
    "Workflow/Admin Regression",
    "Reports/Export Regression",
    "Permission/Security Regression",
    "Mobile/Cross-Browser Regression",
    "CI-CD/Environment Readiness",
    "Defect Retest / Final Signoff",
}

ALLOWED_PRIORITIES = {"P0", "P1", "P2"}
ALLOWED_STATUSES = {"Not Run", "Pass", "Fail", "Blocked", "Setup Required", "Not Applicable"}
ALLOWED_SUBMISSION_STATUSES = {"Not Submitted", "Submitted", "Reviewed", "Blocked", "Setup Required", "Not Applicable"}
EXPECTED_DAY_POINTS = {
    "Day 1": 18.0,
    "Day 2": 13.0,
    "Day 3": 15.0,
    "Day 4": 11.0,
    "Day 5": 9.0,
    "Day 6": 19.0,
    "Day 7": 3.0,
}
EXPECTED_DAY_DATES = {
    "Day 1": "2026-06-10",
    "Day 2": "2026-06-11",
    "Day 3": "2026-06-12",
    "Day 4": "2026-06-13",
    "Day 5": "2026-06-14",
    "Day 6": "2026-06-15",
    "Day 7": "2026-06-16",
}
FORBIDDEN_TERMS = [
    "78 story points",
    "normal teams",
    "normal team",
    "we do 78",
]
CLIENT_SECRET_TERMS = [
    "LocalDev@123",
    "real password",
    "production password",
]
FUTURE_DOMAIN_TERMS = [
    "hawkaii.in",
    "api.hawkaii.in",
    "qa.hawkaii.in",
    "qa-api.hawkaii.in",
    "dev.hawkaii.in",
    "dev-api.hawkaii.in",
]


def col_to_idx(ref: str) -> int:
    letters = "".join(ch for ch in ref if ch.isalpha())
    value = 0
    for ch in letters:
        value = value * 26 + (ord(ch.upper()) - 64)
    return value - 1


def read_with_xml(path: Path) -> tuple[dict[str, list[list[str]]], str]:
    ns = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }
    with ZipFile(path) as zf:
        shared: list[str] = []
        if "xl/sharedStrings.xml" in zf.namelist():
            shared_root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
            for si in shared_root.findall("main:si", ns):
                shared.append("".join(t.text or "" for t in si.findall(".//main:t", ns)))

        workbook = ET.fromstring(zf.read("xl/workbook.xml"))
        rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rel_map = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}
        sheets: dict[str, list[list[str]]] = {}
        sheet_nodes = workbook.find("main:sheets", ns)
        for sheet in sheet_nodes if sheet_nodes is not None else []:
            name = sheet.attrib["name"]
            rel_id = sheet.attrib[f"{{{ns['rel']}}}id"]
            target = rel_map[rel_id]
            xml = ET.fromstring(zf.read(f"xl/{target}"))
            rows: list[list[str]] = []
            for row in xml.findall(".//main:sheetData/main:row", ns):
                values: list[str] = []
                for cell in row.findall("main:c", ns):
                    idx = col_to_idx(cell.attrib.get("r", "A1"))
                    while len(values) <= idx:
                        values.append("")
                    cell_type = cell.attrib.get("t")
                    value = ""
                    if cell_type == "inlineStr":
                        value = "".join(t.text or "" for t in cell.findall(".//main:t", ns))
                    else:
                        v = cell.find("main:v", ns)
                        if v is not None and v.text is not None:
                            value = shared[int(v.text)] if cell_type == "s" else v.text
                    values[idx] = value
                rows.append(values)
            sheets[name] = rows
    return sheets, "XML fallback"


def read_workbook(path: Path) -> tuple[dict[str, list[list[Any]]], str, str]:
    try:
        from openpyxl import load_workbook  # type: ignore

        wb = load_workbook(path, data_only=False, read_only=True)
        sheets: dict[str, list[list[Any]]] = {}
        for ws in wb.worksheets:
            rows: list[list[Any]] = []
            for row in ws.iter_rows(values_only=True):
                values = ["" if cell is None else cell for cell in row]
                while values and values[-1] == "":
                    values.pop()
                rows.append(values)
            sheets[ws.title] = rows
        return sheets, "openpyxl", "Pass"
    except Exception as exc:
        sheets, reader = read_with_xml(path)
        return sheets, reader, f"Not run with openpyxl: {exc}"


def dict_rows(rows: list[list[Any]]) -> list[dict[str, str]]:
    if not rows:
        return []
    headers = [str(cell).strip() for cell in rows[0]]
    result: list[dict[str, str]] = []
    for row in rows[1:]:
        if not any(str(cell).strip() for cell in row):
            continue
        record: dict[str, str] = {}
        for index, header in enumerate(headers):
            record[header] = str(row[index]).strip() if index < len(row) else ""
        result.append(record)
    return result


def parse_date(value: str) -> datetime | None:
    try:
        return datetime.strptime(value, "%Y-%m-%d")
    except ValueError:
        return None


def parse_due(value: str) -> datetime | None:
    try:
        return datetime.strptime(value, "%Y-%m-%d %H:%M IST")
    except ValueError:
        return None


def add_check(condition: bool, errors: list[str], message: str) -> None:
    if not condition:
        errors.append(message)


def priority_pass_rate(rows: list[dict[str, str]], priority: str) -> str:
    relevant = [row for row in rows if row.get("Priority") == priority]
    if not relevant:
        return "0%"
    passed = sum(1 for row in relevant if row.get("Status") == "Pass")
    return f"{passed / len(relevant) * 100:.1f}%"


def main() -> int:
    errors: list[str] = []
    warnings: list[str] = []

    if not WORKBOOK.exists():
        errors.append(f"Missing workbook: {WORKBOOK}")
        sheets: dict[str, list[list[Any]]] = {}
        reader = "none"
        openpyxl_status = "Fail"
    else:
        sheets, reader, openpyxl_status = read_workbook(WORKBOOK)

    for sheet in REQUIRED_SHEETS:
        add_check(sheet in sheets, errors, f"Missing required sheet: {sheet}")

    all_rows: list[dict[str, str]] = []
    detailed_rows: list[dict[str, str]] = []
    ids_by_sheet: dict[str, set[str]] = {}
    priority_counts: Counter[str] = Counter()
    status_counts: Counter[str] = Counter()
    day_points: defaultdict[str, float] = defaultdict(float)
    day_cases: Counter[str] = Counter()
    blockers: list[str] = []
    test_ids: list[str] = []

    for sheet_name in ["Main Test Cases", *EXECUTION_SHEETS]:
        rows = sheets.get(sheet_name, [])
        if not rows:
            errors.append(f"{sheet_name} has no rows")
            continue
        headers = [str(cell).strip() for cell in rows[0]]
        for column in CASE_COLUMNS:
            if column not in headers:
                errors.append(f"{sheet_name} missing column: {column}")

        sheet_ids: set[str] = set()
        for record in dict_rows(rows):
            test_id = record.get("Test Case ID", "")
            if not test_id:
                errors.append(f"{sheet_name} row missing Test Case ID")
                continue
            sheet_ids.add(test_id)
            if sheet_name != "Main Test Cases":
                test_ids.append(test_id)
                detailed_rows.append(record)
                priority_counts[record.get("Priority", "")] += 1
                status_counts[record.get("Status", "")] += 1
                all_rows.append(record)
                day = record.get("Execution Day", "")
                day_cases[day] += 1
                try:
                    day_points[day] += float(record.get("Story Points", ""))
                except ValueError:
                    errors.append(f"{test_id} has invalid Story Points: {record.get('Story Points', '')}")

            for column in CASE_COLUMNS:
                if column in {"Actual Result", "Evidence Link / Screenshot Ref", "Defect ID", "Retest Due Date"}:
                    continue
                if column == "Blocked Reason" and record.get("Status") not in {"Blocked", "Setup Required"}:
                    continue
                if not record.get(column, "").strip():
                    errors.append(f"{test_id} missing {column}")

            priority = record.get("Priority", "")
            if priority not in ALLOWED_PRIORITIES:
                errors.append(f"{test_id} has unsupported priority: {priority}")

            status = record.get("Status", "")
            if status not in ALLOWED_STATUSES:
                errors.append(f"{test_id} has unsupported status: {status}")

            submission_status = record.get("Submission Status", "")
            if submission_status not in ALLOWED_SUBMISSION_STATUSES:
                errors.append(f"{test_id} has unsupported Submission Status: {submission_status}")

            lane = record.get("Execution Lane", "")
            if lane not in LANES:
                errors.append(f"{test_id} has unsupported Execution Lane: {lane}")

            day = record.get("Execution Day", "")
            if day not in EXPECTED_DAY_DATES:
                errors.append(f"{test_id} has unsupported Execution Day: {day}")
            else:
                expected_date = EXPECTED_DAY_DATES[day]
                if record.get("Planned Start Date") != expected_date:
                    errors.append(f"{test_id} Planned Start Date must be {expected_date}")
                if record.get("Planned End Date") != expected_date:
                    errors.append(f"{test_id} Planned End Date must be {expected_date}")
                expected_due = f"{expected_date} {'18:00 IST' if day == 'Day 7' else '20:00 IST'}"
                if record.get("Daily Submission Due") != expected_due:
                    errors.append(f"{test_id} Daily Submission Due must be {expected_due}")

            planned_end = parse_date(record.get("Planned End Date", ""))
            final_due = parse_due(record.get("Final Submission Due", ""))
            if planned_end is None:
                errors.append(f"{test_id} has invalid Planned End Date")
            if final_due is None:
                errors.append(f"{test_id} has invalid Final Submission Due")
            elif final_due > parse_due(FINAL_SUBMISSION):  # type: ignore[operator]
                errors.append(f"{test_id} final submission due is after final QA submission")

            if status in {"Blocked", "Setup Required"}:
                reason = record.get("Blocked Reason", "")
                if not reason:
                    errors.append(f"{test_id} is {status} but has no blocked reason")
                elif sheet_name != "Main Test Cases":
                    blockers.append(f"{test_id}: {reason}")

            combined = " ".join(record.values()).lower()
            if any(term in combined for term in FUTURE_DOMAIN_TERMS):
                if status not in {"Blocked", "Setup Required"}:
                    errors.append(f"{test_id} references future domains but is not Blocked/Setup Required")

        ids_by_sheet[sheet_name] = sheet_ids

    main_ids = ids_by_sheet.get("Main Test Cases", set())
    detailed_ids = set(test_ids)
    if main_ids and detailed_ids and main_ids != detailed_ids:
        missing_from_main = sorted(detailed_ids - main_ids)
        missing_from_details = sorted(main_ids - detailed_ids)
        if missing_from_main:
            errors.append(f"Main Test Cases missing IDs: {', '.join(missing_from_main)}")
        if missing_from_details:
            errors.append(f"Detailed sheets missing IDs from Main Test Cases: {', '.join(missing_from_details)}")

    duplicate_ids = sorted({test_id for test_id in test_ids if test_ids.count(test_id) > 1})
    if duplicate_ids:
        errors.append(f"Duplicate test case IDs: {', '.join(duplicate_ids)}")

    for priority in ["P0", "P1", "P2"]:
        if priority_counts[priority] == 0:
            errors.append(f"Missing priority group {priority}")

    for day, expected in EXPECTED_DAY_POINTS.items():
        actual = day_points.get(day, 0.0)
        if abs(actual - expected) > 0.001:
            errors.append(f"{day} expected {expected:g} story points, found {actual:g}")

    for sheet_name in ["Role Permission Matrix", "Business Rule Traceability"]:
        rows = sheets.get(sheet_name, [])
        text = "\n".join("\t".join(str(cell) for cell in row) for row in rows)
        for required in ["Execution Day", "Execution Lane", "Submission Due"]:
            if required not in text:
                errors.append(f"{sheet_name} missing timeline marker: {required}")

    signoff_text = "\n".join("\t".join(str(cell) for cell in row) for row in sheets.get("Signoff", []))
    for required in [
        "Planned QA start date",
        "Planned completion date",
        "Final submission date/time",
        "P0 pass rate",
        "P1 pass rate",
        "Release recommendation",
    ]:
        if required not in signoff_text:
            errors.append(f"Signoff missing: {required}")

    workbook_text = "\n".join(
        "\n".join("\t".join(str(cell) for cell in row) for row in rows) for rows in sheets.values()
    )
    for term in FORBIDDEN_TERMS:
        if term.lower() in workbook_text.lower():
            errors.append(f"Workbook contains forbidden wording: {term}")

    client_doc = QA / "TESTING_CHECKLIST_CLIENT.md"
    client_text = client_doc.read_text(encoding="utf-8") if client_doc.exists() else ""
    for term in CLIENT_SECRET_TERMS:
        if term.lower() in client_text.lower() or term.lower() in workbook_text.lower():
            errors.append(f"Client-facing output exposes forbidden secret wording: {term}")

    for doc in [
        QA / "TESTING_CHECKLIST_INTERNAL.md",
        QA / "TESTING_CHECKLIST_CLIENT.md",
        QA / "TESTER_RUN_BOOK.md",
        QA / "QA_TIMELINE_EXECUTION_PLAN.md",
        QA / "QA_DAILY_SUBMISSION_PLAN.md",
    ]:
        if not doc.exists():
            errors.append(f"Missing document: {doc.relative_to(ROOT)}")
            continue
        text = doc.read_text(encoding="utf-8")
        for term in FORBIDDEN_TERMS:
            if term.lower() in text.lower():
                errors.append(f"{doc.relative_to(ROOT)} contains forbidden wording: {term}")

    if "python" not in openpyxl_status.lower() and openpyxl_status != "Pass":
        warnings.append(openpyxl_status)
    elif openpyxl_status != "Pass":
        warnings.append(openpyxl_status)

    result = "Pass" if not errors else "Fail"
    total_cases = len(test_ids)
    lines = [
        "# QA Timeline Validation Log",
        "",
        "## Files Changed",
        *[f"- `{path}`" for path in FILES_CHANGED],
        "",
        "## Timeline",
        f"- Start date used: `{START_DATE}`",
        f"- Start time used: `{START_TIME}`",
        f"- Completion date used: `{COMPLETION_DATE}`",
        f"- Final submission date/time used: `{FINAL_SUBMISSION}`",
        "",
        "## Sheet Validation Result",
        f"- Workbook: `{WORKBOOK}`",
        f"- Reader used by validator: `{reader}`",
        f"- Openpyxl load result: `{openpyxl_status}`",
        f"- Result: `{result}`",
        "",
        "## Counts",
        f"- Total test cases: {total_cases}",
        f"- P0: {priority_counts['P0']}",
        f"- P1: {priority_counts['P1']}",
        f"- P2: {priority_counts['P2']}",
        f"- Completed: {status_counts['Pass'] + status_counts['Fail'] + status_counts['Not Applicable']}",
        f"- Passed: {status_counts['Pass']}",
        f"- Failed: {status_counts['Fail']}",
        f"- Blocked: {status_counts['Blocked']}",
        f"- Setup Required: {status_counts['Setup Required']}",
        f"- Not run: {status_counts['Not Run']}",
        f"- P0 pass rate: {priority_pass_rate(detailed_rows, 'P0')}",
        f"- P1 pass rate: {priority_pass_rate(detailed_rows, 'P1')}",
        "",
        "## Story Points By Day",
        *[f"- {day}: {day_points.get(day, 0):g} SP" for day in sorted(EXPECTED_DAY_POINTS)],
        "",
        "## Test Cases By Day",
        *[f"- {day}: {day_cases.get(day, 0)} cases" for day in sorted(EXPECTED_DAY_POINTS)],
        "",
        "## Blockers",
        *([f"- {blocker}" for blocker in blockers] or ["- None"]),
        "",
        "## Commands Run",
        "- `python3 -m pip install --target /private/tmp/hawkaii-openpyxl openpyxl`",
        "- `PYTHONPATH=/private/tmp/hawkaii-openpyxl python3 - <<'PY' ... load_workbook('qa/TESTING_TEST_CASES.xlsx') ... PY`",
        "- `PYTHONPATH=/private/tmp/hawkaii-openpyxl python3 qa/scripts/validate_testing_workbook.py`",
        "",
        "## Errors",
        *([f"- {error}" for error in errors] or ["- None"]),
        "",
        "## Warnings",
        *([f"- {warning}" for warning in warnings] or ["- None"]),
        "",
        "## Manual Follow-Up Needed",
        "- Configure isolated future domains/services before changing future-domain readiness rows from `Setup Required`.",
        "- Document a safe production test tenant and reset process before allowing production create/update/delete tests.",
    ]
    TIMELINE_LOG.write_text("\n".join(lines) + "\n", encoding="utf-8")

    legacy_lines = [
        "# QA Sheet Validation Log",
        "",
        f"Workbook: `{WORKBOOK}`",
        f"Result: {result}",
        "",
        "## Counts",
        f"- Test cases: {total_cases}",
        f"- P0: {priority_counts['P0']}",
        f"- P1: {priority_counts['P1']}",
        f"- P2: {priority_counts['P2']}",
        f"- Day totals: {', '.join(f'{day}: {day_points.get(day, 0):g} SP' for day in sorted(EXPECTED_DAY_POINTS))}",
        "",
        "## Errors",
        *([f"- {error}" for error in errors] or ["- None"]),
        "",
        "## Warnings",
        *([f"- {warning}" for warning in warnings] or ["- None"]),
    ]
    LEGACY_LOG.write_text("\n".join(legacy_lines) + "\n", encoding="utf-8")

    print("\n".join(lines))
    return 1 if errors else 0


if __name__ == "__main__":
    raise SystemExit(main())
